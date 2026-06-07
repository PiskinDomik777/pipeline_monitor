from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User, Pipeline, Defect
from analytics import DefectAnalyzer
import pandas as pd
import os
from werkzeug.utils import secure_filename
from werkzeug.utils import redirect
from werkzeug.security import generate_password_hash
from datetime import datetime
from flask import send_file

# Импорт для админ-панели
from flask_admin import Admin
from flask_admin.contrib.sqla import ModelView
from flask_admin.base import AdminIndexView 
from flask_admin.menu import MenuLink 

app = Flask(__name__)

# Конфигурация
app.config['SECRET_KEY'] = 'gazprom-super-secret-key-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://ISPr25-24_LinkovNI:ISPr25-24_LinkovNI@cfif31.ru:3306/ISPr25-24_LinkovNI_pipeline_db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# Логин
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Пожалуйста, войдите в систему'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Админ панель
class AdminModelView(ModelView):
    
    def is_accessible(self):
        return current_user.is_authenticated and current_user.is_admin()
    
    def inaccessible_callback(self, name, **kwargs):
        return redirect('/login')


admin = Admin(app, name='ВТД Админ')

admin.add_view(AdminModelView(Pipeline, db))
admin.add_view(AdminModelView(Defect, db))
admin.add_view(AdminModelView(User, db))

from flask_admin.menu import MenuLink
admin.add_link(MenuLink(name='← На главный сайт', url='/'))

# Создание таблиц и тестовых данных
with app.app_context():
    db.create_all()
    
    if User.query.count() == 0:
        admin_user = User(username='admin', email='admin@gazprom.ru', role='admin')
        admin_user.set_password('admin123')
        db.session.add(admin_user)
        
        normal_user = User(username='user', email='user@gazprom.ru', role='user')
        normal_user.set_password('user123')
        db.session.add(normal_user)
        
        db.session.commit()
        print("✅ Созданы тестовые пользователи:")
        print("   Админ: admin / admin123")
        print("   Пользователь: user / user123")
    
    if Pipeline.query.count() == 0:
        pipeline1 = Pipeline(name="Газопровод 'Уренгой-Помары-Ужгород'", length_km=4451)
        pipeline2 = Pipeline(name="Газопровод 'Ямал-Европа'", length_km=2000)
        db.session.add(pipeline1)
        db.session.add(pipeline2)
        db.session.commit()
        print("✅ Созданы тестовые трубопроводы")

# Авторизация
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            flash(f'Добро пожаловать, {user.username}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Неверное имя пользователя или пароль', 'danger')
    
    return render_template('login.html')

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """Личный кабинет пользователя"""
    from werkzeug.utils import secure_filename
    from PIL import Image
    
    if request.method == 'POST':
        new_username = request.form.get('username')
        if new_username and new_username != current_user.username:
            existing_user = User.query.filter_by(username=new_username).first()
            if existing_user:
                flash('Пользователь с таким именем уже существует', 'danger')
            else:
                current_user.username = new_username
                flash('Имя пользователя обновлено', 'success')
        
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if old_password and new_password:
            if not current_user.check_password(old_password):
                flash('Неверный текущий пароль', 'danger')
            elif new_password != confirm_password:
                flash('Новый пароль и подтверждение не совпадают', 'danger')
            elif len(new_password) < 4:
                flash('Пароль должен содержать минимум 4 символа', 'danger')
            else:
                current_user.set_password(new_password)
                flash('Пароль успешно изменён', 'success')
        
        db.session.commit()
        return redirect(url_for('profile'))
    
    return render_template('profile.html', user=current_user, is_admin=current_user.is_admin())

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    pipelines = Pipeline.query.all()
    defects_count = Defect.query.count()
    risk_stats = {
        'critical': Defect.query.filter_by(risk_level='КРИТИЧЕСКИЙ').count(),
        'high': Defect.query.filter_by(risk_level='ВЫСОКИЙ').count(),
        'medium': Defect.query.filter_by(risk_level='СРЕДНИЙ').count(),
        'low': Defect.query.filter_by(risk_level='НИЗКИЙ').count()
    }
    return render_template('index.html', 
                         pipelines=pipelines,
                         defects_count=defects_count,
                         risk_stats=risk_stats,
                         is_admin=current_user.is_admin(),
                         is_user=current_user.role == 'user',
                         is_viewer=current_user.role == 'viewer')

# ==================== АДМИН-ПАНЕЛЬ (КАСТОМНАЯ) ====================

@app.route('/admin-panel')
@login_required
def admin_panel():
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    pipelines_count = Pipeline.query.count()
    defects_count = Defect.query.count()
    users_count = User.query.count()
    users = User.query.all()
    risk_stats = {
        'critical': Defect.query.filter_by(risk_level='КРИТИЧЕСКИЙ').count(),
        'high': Defect.query.filter_by(risk_level='ВЫСОКИЙ').count(),
        'medium': Defect.query.filter_by(risk_level='СРЕДНИЙ').count(),
        'low': Defect.query.filter_by(risk_level='НИЗКИЙ').count()
    }
    
    return render_template('admin_panel.html',
                         pipelines_count=pipelines_count,
                         defects_count=defects_count,
                         users_count=users_count,
                         users=users,
                         risk_stats=risk_stats)

@app.route('/admin-panel/add-user', methods=['POST'])
@login_required
def admin_panel_add_user():
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role')
    
    existing_user = User.query.filter_by(username=username).first()
    if existing_user:
        flash('Пользователь с таким логином уже существует', 'danger')
        return redirect(url_for('admin_panel'))
    
    new_user = User(username=username, email=email, role=role)
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    
    flash(f'Пользователь {username} создан', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin-panel/delete-user/<int:user_id>')
@login_required
def admin_panel_delete_user(user_id):
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('Нельзя удалить себя', 'danger')
        return redirect(url_for('admin_panel'))
    
    db.session.delete(user)
    db.session.commit()
    flash(f'Пользователь удален', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin-panel/edit-user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def admin_panel_edit_user(user_id):
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        user.username = request.form.get('username')
        user.email = request.form.get('email')
        user.role = request.form.get('role')
        
        new_password = request.form.get('password')
        if new_password:
            user.set_password(new_password)
        
        db.session.commit()
        flash('Пользователь обновлен', 'success')
        return redirect(url_for('admin_panel'))
    
    return render_template('admin_panel_edit.html', user=user)


@app.route('/upload', methods=['GET', 'POST'])
@login_required
def upload_data():
    if not current_user.is_admin():
        flash('Доступ запрещен. Только администраторы могут загружать данные.', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Нет файла', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Файл не выбран', 'danger')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            df = pd.read_csv(filepath, encoding='utf-8')
            pipeline_id = request.form.get('pipeline_id', 1)
            
            count = 0
            for _, row in df.iterrows():
                defect_type = row.get('defect_type', 'коррозия')
                depth_mm = float(row.get('depth_mm', 0))
                length_mm = float(row.get('length_mm', 0))
                width_mm = float(row.get('width_mm', 0)) if 'width_mm' in row else None
                km = float(row.get('km', 0))
                
                risk_level, risk_score, recommendation = DefectAnalyzer.analyze_defect(
                    defect_type, depth_mm, length_mm, width_mm
                )
                
                defect = Defect(
                    pipeline_id=pipeline_id,
                    km=km,
                    defect_type=defect_type,
                    depth_mm=depth_mm,
                    length_mm=length_mm,
                    width_mm=width_mm,
                    risk_level=risk_level,
                    risk_score=risk_score,
                    recommendation=recommendation
                )
                db.session.add(defect)
                count += 1
            
            db.session.commit()
            flash(f'Успешно загружено {count} дефектов!', 'success')
            return redirect(url_for('defects_list'))
    
    pipelines = Pipeline.query.all()
    return render_template('upload.html', pipelines=pipelines, is_admin=current_user.is_admin())

@app.route('/defects')
@login_required
def defects_list():
    defects = Defect.query.order_by(Defect.risk_score.desc()).all()
    return render_template('defects.html', 
                         defects=defects, 
                         is_admin=current_user.is_admin(),
                         is_user=current_user.role == 'user',
                         is_viewer=current_user.role == 'viewer')

@app.route('/defects/delete/<int:defect_id>')
@login_required
def delete_defect(defect_id):
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('defects_list'))
    
    defect = Defect.query.get_or_404(defect_id)
    db.session.delete(defect)
    db.session.commit()
    flash('Дефект удален', 'success')
    return redirect(url_for('defects_list'))

@app.route('/defects/edit/<int:defect_id>', methods=['GET', 'POST'])
@login_required
def edit_defect(defect_id):
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('defects_list'))
    
    defect = Defect.query.get_or_404(defect_id)
    
    if request.method == 'POST':
        defect.km = float(request.form.get('km'))
        defect.defect_type = request.form.get('defect_type')
        defect.depth_mm = float(request.form.get('depth_mm'))
        defect.length_mm = float(request.form.get('length_mm'))
        defect.risk_level = request.form.get('risk_level')
        defect.recommendation = request.form.get('recommendation')
        
        db.session.commit()
        flash('Дефект обновлен', 'success')
        return redirect(url_for('defects_list'))
    
    return render_template('edit_defect.html', 
                         defect=defect, 
                         is_admin=current_user.is_admin(),
                         is_user=current_user.role == 'user',
                         is_viewer=current_user.role == 'viewer')

@app.route('/pipelines')
@login_required
def pipelines_list():
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    pipelines = Pipeline.query.all()
    return render_template('pipelines.html', pipelines=pipelines, is_admin=current_user.is_admin())

@app.route('/pipelines/add', methods=['GET', 'POST'])
@login_required
def add_pipeline():
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        length_km = float(request.form.get('length_km'))
        
        pipeline = Pipeline(name=name, length_km=length_km)
        db.session.add(pipeline)
        db.session.commit()
        flash('Трубопровод добавлен', 'success')
        return redirect(url_for('pipelines_list'))
    
    return render_template('add_pipeline.html', is_admin=current_user.is_admin())

@app.route('/pipelines/delete/<int:pipeline_id>')
@login_required
def delete_pipeline(pipeline_id):
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    pipeline = Pipeline.query.get_or_404(pipeline_id)
    db.session.delete(pipeline)
    db.session.commit()
    flash('Трубопровод удален', 'success')
    return redirect(url_for('pipelines_list'))

# УПРАВЛЕНИЕ ПОЛЬЗОВАТЕЛЯМИ (АДМИН) 
@app.route('/admin/users')
@login_required
def admin_users():
    """Страница управления пользователями (только для администратора)"""
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    users = User.query.all()
    return render_template('admin_users.html', users=users, is_admin=current_user.is_admin())

@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
def admin_add_user():
    """Добавление нового пользователя (только для администратора)"""
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        
        # Проверка на существующего пользователя
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Пользователь с таким логином уже существует', 'danger')
            return redirect(request.url)
        
        existing_email = User.query.filter_by(email=email).first()
        if existing_email:
            flash('Пользователь с таким email уже существует', 'danger')
            return redirect(request.url)
        
        new_user = User(username=username, email=email, role=role)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        
        flash(f'Пользователь {username} успешно создан', 'success')
        return redirect(url_for('admin_users'))
    
    return render_template('admin_add_user.html', is_admin=current_user.is_admin())

@app.route('/admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def admin_edit_user(user_id):
    """Редактирование пользователя (только для администратора)"""
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        user.username = request.form.get('username')
        user.email = request.form.get('email')
        user.role = request.form.get('role')
        
        new_password = request.form.get('password')
        if new_password:
            user.set_password(new_password)
        
        db.session.commit()
        flash(f'Пользователь {user.username} обновлен', 'success')
        return redirect(url_for('admin_users'))
    
    return render_template('admin_edit_user.html', user=user, is_admin=current_user.is_admin())

@app.route('/admin/users/delete/<int:user_id>')
@login_required
def admin_delete_user(user_id):
    """Удаление пользователя (только для администратора)"""
    if not current_user.is_admin():
        flash('Доступ запрещен', 'danger')
        return redirect(url_for('index'))
    
    user = User.query.get_or_404(user_id)
    
    # Нельзя удалить самого себя
    if user.id == current_user.id:
        flash('Нельзя удалить свою учетную запись', 'danger')
        return redirect(url_for('admin_users'))
    
    db.session.delete(user)
    db.session.commit()
    flash(f'Пользователь {user.username} удален', 'success')
    return redirect(url_for('admin_users'))

@app.route('/report')
@login_required
def report():
    defects = Defect.query.order_by(Defect.risk_score.desc()).all()
    pipeline = Pipeline.query.first()
    return render_template('report.html', 
                         defects=defects, 
                         pipeline=pipeline, 
                         is_admin=current_user.is_admin(),
                         is_user=current_user.role == 'user',
                         is_viewer=current_user.role == 'viewer')

@app.route('/export/excel')
@login_required
def export_excel():
    """Экспорт всех дефектов в Excel-файл"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    defects = Defect.query.order_by(Defect.risk_score.desc()).all()
    pipeline = Pipeline.query.first()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Дефекты ВТД"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f'ПАО "Газпром" - Технический отчёт по внутритрубной диагностике'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Объект: {pipeline.name if pipeline else "Газопровод"}'
    ws['A2'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A3:G3')
    ws['A3'] = f'Дата формирования: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws['A3'].alignment = Alignment(horizontal="center")
    
    headers = ['№', 'Километр', 'Тип дефекта', 'Глубина (мм)', 'Длина (мм)', 'Уровень риска', 'Рекомендация']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for idx, defect in enumerate(defects, 1):
        row = 5 + idx
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=defect.km).border = thin_border
        ws.cell(row=row, column=3, value=defect.defect_type).border = thin_border
        ws.cell(row=row, column=4, value=defect.depth_mm).border = thin_border
        ws.cell(row=row, column=5, value=defect.length_mm).border = thin_border
        ws.cell(row=row, column=6, value=defect.risk_level).border = thin_border
        ws.cell(row=row, column=7, value=defect.recommendation).border = thin_border
        
        risk_cell = ws.cell(row=row, column=6)
        if defect.risk_level == 'КРИТИЧЕСКИЙ':
            risk_cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            risk_cell.font = Font(color="FFFFFF", bold=True)
        elif defect.risk_level == 'ВЫСОКИЙ':
            risk_cell.fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
            risk_cell.font = Font(color="FFFFFF", bold=True)
        elif defect.risk_level == 'СРЕДНИЙ':
            risk_cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        elif defect.risk_level == 'НИЗКИЙ':
            risk_cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            risk_cell.font = Font(color="FFFFFF", bold=True)
    
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'otchet_vtd_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

from datetime import datetime
from flask import send_file

@app.route('/api/defects')
@login_required
def api_defects():
    defects = Defect.query.all()
    return jsonify([defect.to_dict() for defect in defects])

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=3000)
    import plotly.graph_objs as go
import plotly.utils
import json

@app.route('/charts_data')
@login_required
def charts_data():
    defects = Defect.query.all()
    
    # Данные для графика
    risk_counts = {
        'КРИТИЧЕСКИЙ': 0,
        'ВЫСОКИЙ': 0,
        'СРЕДНИЙ': 0,
        'НИЗКИЙ': 0
    }
    type_counts = {}
    km_data = []
    
    for d in defects:
        risk_counts[d.risk_level] = risk_counts.get(d.risk_level, 0) + 1
        type_counts[d.defect_type] = type_counts.get(d.defect_type, 0) + 1
        km_data.append({'km': d.km, 'risk': d.risk_level, 'type': d.defect_type})
    
    return jsonify({
        'risk_counts': risk_counts,
        'type_counts': type_counts,
        'km_data': km_data
    })
